import requests
from openpyxl import load_workbook
import regex
import json
import time

#将原先的文本转化成更利于搜索的文本，
#输入是原先文本在excel表中的索引
#无返回值
def processPlainText(index):

    row = list(sheet.rows)[index]
    oriAddress = row[3].value
    obj["oriAddress"] = row[3].value
    searchReg =  regex.search(r'(?<=[\(,（]\s*)\w*(?=方)',oriAddress)
    if searchReg != None:
        direction =searchReg.group()
        obj["address"] = oriAddress.split("服务区")[0]+"服务区"+direction +"方向"
    else:
        obj["address"] = oriAddress.split("服务区")[0]+"服务区"
#向api发起请求，并传输相关的地址以获得相关返回值，包括精度，纬度，省，市，区，并将其保存在地址中
#无输入值
#无返回值
def gettingLocAndSave():
    link = "https://restapi.amap.com/v3/place/text"
    params ={"key":"62a9b8d08aa2b7ec94e5952643e0a885","keywords":obj["address"],"type":"充电站"}
    res = requests.get(link,params = params)
    if len(json.loads(res.text)["pois"]) !=0:
        lat = json.loads(res.text)["pois"][0]["location"].split(",")[0]
        lon = json.loads(res.text)["pois"][0]["location"].split(",")[1]
        pname = json.loads(res.text)["pois"][0]["pname"]
        cityname = json.loads(res.text)["pois"][0]["cityname"]
        adname = json.loads(res.text)["pois"][0]["adname"]
        row = [obj["oriAddress"],obj["address"],lat,lon,pname,cityname,adname]
    else:
        row = [obj["oriAddress"]]
    print(row)
    write_sheet.append(row)
    write_wb.save("./doc/result.xlsx")
#主函数，主要的作用是加载有源数据的ServiceArea.xlsx以及保存数据的result.xlsx
if __name__ =='__main__':
    wb = load_workbook("./doc/ServiceArea.xlsx")
    write_wb = load_workbook("./doc/result.xlsx")
    sheet = wb['Sheet1']
    write_sheet = write_wb['Sheet1']
    for i in range(246,1523):#确定源数据所在的范围，即所在的列，第一位数-1
        obj={}
        processPlainText(i)
        gettingLocAndSave()
